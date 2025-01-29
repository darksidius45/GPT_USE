from googlesearch import search
# import re
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.options import Options
import time
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import WebDriverException
import openpyxl
import json
import subprocess
import requests
import random
from bs4 import BeautifulSoup
# import zipfile
# from config import promt
import pandas as pd
from urllib.parse import urlparse, urljoin
import tempfile
import os

PROXY_PORT = 5500
PROXY_USER = "5vWk59"
PROXY_PASSWORD = "tXQt5fnHW9"

def get_random_proxy():
    proxies = [
        "45.11.21.124",
        "213.226.101.147",
        "94.158.190.16",
        "109.248.128.64",
        "188.130.218.160",
        "46.8.15.51",
        "46.8.23.126",
        "188.130.129.246",
        "188.130.128.66",
        "46.8.106.76",
        "46.8.111.80",
        "109.248.142.228",
        "46.8.222.147",
        "109.248.204.170"
    ]
    return random.choice(proxies)




def find_website(name):
    query = f"{name} site"
    proxy = get_random_proxy()
    proxy_url = f"http://{PROXY_USER}:{PROXY_PASSWORD}@{proxy}:{PROXY_PORT}"

    try:
        for result in search(query, num_results=10):
            if all(exclusion not in result for exclusion in ['youtube.com', 'facebook.com', 'wikipedia.org', 'linkedin.com']):
                return result
    except requests.exceptions.HTTPError as e:
        print(f"HTTP error occurred: {e}")
    return None



def get_site_text(website):
    chrome_options = Options()
    chrome_options.add_argument("user-agent=Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/605.1.15 (KHTML, like Gecko) Version/14.0.3 Safari/605.1.15")
    service = Service("C:/webdrivers/chromedriver.exe")
    driver = webdriver.Chrome(service=service, options=chrome_options)
    original_domain = urlparse(website).netloc
    keywords = [
        "about", "product", "solution",  # English
        "über", "produkt", "lösung",  # German
        "acerca de", "producto", "solución",  # Spanish
        "à propos", "produit", "solution",  # French
        "circa", "prodotto", "soluzione",  # Italian
        "sobre", "produto", "solução",  # Portuguese
        "关于", "产品", "解决方案",  # Chinese
        "о", "продукт", "решение",  # Russian
        "について", "製品", "ソリューション",  # Japanese
        "hakkında", "ürün", "çözüm",  # Turkish
        "حول", "منتج", "حل"  # Arabic
    ]
    all_texts = []

    try:
        try:  # Set a longer page load timeout
            driver.get(website)
            
        except WebDriverException as e:
            print(f"WebDriverException occurred while accessing {website}: {e}")
            driver.quit()
            return None
        
        page_source = driver.page_source
        soup = BeautifulSoup(page_source, "html.parser")
        all_texts.append(soup.get_text(separator=" ", strip=True))

        hrefs_set = set()
        clickable_elements = set()
        for element in soup.find_all(['a', 'button', 'input', 'area']):
            if element.has_attr('href'):
                href = element['href']
                if href not in hrefs_set and original_domain in urlparse(href).netloc and any(keyword in href for keyword in keywords):
                    clickable_elements.add(element['href'])
        for i in clickable_elements:
            print(i, end="\n")
        for element in clickable_elements:
            new_link = urljoin(website, element)
            driver.get(new_link)
            WebDriverWait(driver, 20).until(EC.presence_of_element_located((By.TAG_NAME, 'body')))  # Wait for the new page to load
            page_source = driver.page_source
            soup = BeautifulSoup(page_source, "html.parser")
            all_texts.append(soup.get_text(separator=" ", strip=True))
    except WebDriverException as e:
        print(f"WebDriverException occurred while accessing {website}: {e}")
    except TimeoutError as e:
        print(f"TimeoutError occurred while accessing {website}: {e}")
    except Exception as e:
        print(f"An unexpected error occurred while accessing {website}: {e}")
    finally:
        driver.quit()

    return " ".join(all_texts)


def gemini_flash(text):
    # Create a temporary file to store the JSON data
    with tempfile.NamedTemporaryFile(delete=False, mode='w', suffix='.json') as temp_file:
        json_data = {
            "contents": [{"role": "user", "parts": [{"text": text}]}]
        }
        json.dump(json_data, temp_file)
        temp_file.flush()  # Ensure the data is written

        # Get the name of the temporary file
        temp_file_name = temp_file.name

    try:
        # Simplified curl command using the temporary file
        curl_command = [
            "curl",
            "https://api.proxyapi.ru/google/v1/models/gemini-1.5-flash:generateContent",
            "-H", "Content-Type: application/json",
            "-H", "Authorization: Bearer sk-UmoxJZ8wJm1DEmcSrwP3Iu9Bk4TGZJ0h",
            "-d", f"@{temp_file_name}"  # Use the temp file
        ]

        result = subprocess.run(curl_command, capture_output=True, text=True, encoding='utf-8')
        if result.returncode == 0:
            response = json.loads(result.stdout)
            text_response = response['candidates'][0]['content']['parts'][0]['text']
            print(text_response)
            return text_response
        else:
            print("Error executing curl command:", result.stderr)
            return None
    finally:
        # Clean up the temporary file
        if os.path.exists(temp_file_name):
            os.remove(temp_file_name)



# Load the existing workbook
df = openpyxl.load_workbook("out.xlsx")
sheet = df.active
new_workbook = openpyxl.load_workbook("new_output2.xlsx")
new_sheet = new_workbook.active

for row in sheet.iter_rows(min_row=894):  # Changed min_row to 7
    name = row[0].value
    website = row[1].value
    site_text = get_site_text(website)
    new_sheet.append([name, site_text])  # Save name and site_text to the new workbook
    new_workbook.save(f"new_output2.xlsx")

# Save the changes to the original Excel file
df.close()
