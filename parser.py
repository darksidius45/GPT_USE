# from googlesearch import search
# import re
# from selenium import webdriver
# from selenium.webdriver.chrome.service import Service
# from selenium.webdriver.common.by import By
# from selenium.webdriver.chrome.options import Options
# import time
# from selenium.webdriver.support.ui import WebDriverWait
# from selenium.webdriver.support import expected_conditions as EC
# from selenium.common.exceptions import WebDriverException

# import json
# import subprocess
# import requests
# import random
# from bs4 import BeautifulSoup
# import zipfile
# from config import promt


# PROXY_PORT = 5500
# PROXY_USER = "5vWk59"
# PROXY_PASSWORD = "tXQt5fnHW9"


# def get_company_website(name, exhibitor_url):
#     try:
#         chrome_options = Options()
#         chrome_options.add_argument("user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/114.0.5735.110 Safari/537.36")
#         service = Service("C:/webdrivers/chromedriver.exe")
#         driver = webdriver.Chrome(service=service, options=chrome_options)
#         try:
#             driver.get(exhibitor_url)
#             time.sleep(3)
#             site_element = driver.page_source
#             soup = BeautifulSoup(site_element, "html.parser")
#             phone = soup.find("span", class_="dib muted").find_next_sibling(text=True).strip()
#             website = soup.find("ul", class_="showcase-web-phone ml0 mb3 list tc").find('a')['href']
#             address = soup.find("p", class_="showcase-address tc").get_text(separator=" ", strip=True)
#             categories = []
#             categories_section = soup.find("div", role="list", class_="grid grid-3-col grid__centered")
#             if categories_section:
#                 category_items = categories_section.find_all("div", role="listitem")
#                 for item in category_items:
#                     category_name = item.find("h2").get_text(strip=True)
#                     categories.append(category_name)
#             description_element = soup.find("p", class_="js-read-more animated")
#             description = description_element.get_text(separator=" ", strip=True) if description_element else "Description not found"
#             print(f"Website: {website}, Phone: {phone}, Address: {address}, Categories: {categories}, Description: {description}")
#             driver.quit()
#             return website, phone, address, categories, description
#         except WebDriverException as e:
#             print(f"WebDriverException occurred: {e}")
#             driver.quit()
#             return "error", "error", "error", "error", "error"

#     except Exception as e:
#         print(e)
#         return "error", "error", "error", "error", "error"

# def check_av_integration_needs_with_gpt(prompt):
#     # chrome_options = Options()
#     # chrome_options.add_argument("user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/114.0.5735.110 Safari/537.36")
#     # service = Service("C:/webdrivers/chromedriver.exe")
#     # driver = webdriver.Chrome(service=service, options=chrome_options)
#     # try:
#     #     driver.get(site)
#     #     time.sleep(3)
#     #     site_element = driver.page_source
#     #     soup = BeautifulSoup(site_element, "html.parser")
#     #     hrefs = []

#     #     for a_tag in soup.find_all('a', href=True):
#     #         href = a_tag['href']
#     #         if not any(exclusion in href for exclusion in ['youtube.com', 'wikipedia.org', 'facebook.com', 'instagram.com']):
#     #             if not any(ext in href for ext in ['.jpg', '.jpeg', '.png', '.gif', '.mp4', '.avi', '.mov']):
#     #                 hrefs.append(href)
#     #     main_page_text = []
#     #     for href in hrefs:
#     #         try:
#     #             driver.get(href)
#     #             time.sleep(3)
#     #             href_page_source = driver.page_source
#     #             href_soup = BeautifulSoup(href_page_source, "html.parser")
#     #             main_page_text.append(href_soup.get_text(separator=" ", strip=True))

#     #         except requests.exceptions.RequestException as e:
#     #             print(f"Error fetching {href}: {e}")

#     #     print(main_page_text)
#     #     driver.quit()
#     # except WebDriverException as e:
#     #     print(main_page_text)
#     #     print(f"WebDriverException occurred: {e}")
#     # finally:
#     #     driver.quit()


#     # prompt = f"tell me about this company: you have this text about it {main_page_text}"
#     # try:
#     #     curl_command = [

#     #     "curl",
#     #     "https://api.proxyapi.ru/openai/v1/chat/completions",
#     #     "-H", "Content-Type: application/json",
#     #     "-H", "Authorization: Bearer sk-UmoxJZ8wJm1DEmcSrwP3Iu9Bk4TGZJ0h",
#     #     "-d", json.dumps({
#     #         "model": "gpt-4o-mini",
#     #         "messages": [{"role": "user", "content": prompt}],
#     #         "stream": False
#     #     })
#     #     ]


#     #     result = subprocess.run(curl_command, capture_output=True, text=True, encoding='utf-8')
#         curl_command_2 = [
#             "curl",
#             "https://api.proxyapi.ru/openai/v1/chat/completions",
#             "-H", "Content-Type: application/json",
#             "-H", "Authorization: Bearer sk-UmoxJZ8wJm1DEmcSrwP3Iu9Bk4TGZJ0h",
#             "-d", json.dumps({
#                 "model": "gpt-4o-mini",
#                 "messages": [
#                     {"role": "user", "content": f"is this company connected with av integration and how, first write yes or no, then write why?{prompt} {promt}"}
#                 ],
#                 "stream": False
#             })
#         ]

#         result_2 = subprocess.run(curl_command_2, capture_output=True, text=True, encoding='utf-8')
#         response = json.loads(result_2.stdout)
#         print(response['choices'][0]['message']['content'])
#         return response['choices'][0]['message']['content']


#     #     if result_2.returncode != 0:
#     #         print("Error executing second curl command:", result_2.stderr)
#     #         driver.quit()
#     #         return "rrrrr", "rrrrr"

#     #     response_2 = json.loads(result_2.stdout)

#     #     if result.returncode != 0:
#     #         print("Error executing curl command:", result.stderr)
#     #         driver.quit()
#     #         return "rrrrr", "rrrrr"

#     #     response = json.loads(result.stdout)
#     #     print(response['choices'][0]['message']['content'])
#     #     print(response_2['choices'][0]['message']['content'])
#     #     return response['choices'][0]['message']['content'], response_2['choices'][0]['message']['content']
#     # except Exception as e:
#     #     print(f"Error checking site {site}: {e}")
#     #     driver.quit()
#     #     return "eror", "eror"


# output_file = r"C:\Users\prive\Desktop\prog\rgb\GPT_USE\output.xlsx"
# file_pars = r"C:\Users\prive\Downloads\pars.txt"
# plugin_file = "proxy_auth_plugin.zip"

# company_names = set()

# with open(file_pars, 'r') as f:
#     content = f.read()

#     soup = BeautifulSoup(content, 'html.parser')
#     for h3_tag in soup.find_all('h3', class_='card-Title break-word f2 mb1 mt0'):
#         a_tag = h3_tag.find('a', class_='bb-0')
#         if a_tag:
#             href = "https://ise2025.mapyourshow.com" + a_tag['href']
#             company_name = a_tag.find('span').text.replace('\n', '')
#             company_names.add((company_name, href))


# for match in company_names:
#     name = match[0]
#     website = match[1]

#     if name and website:
#         website, phone, address, categories, description = get_company_website(name, website)
#         # Open the output.xlsx file and save all data
#         try:
#             workbook = openpyxl.load_workbook("out.xlsx")
#             sheet = workbook.active

#             # Find the next empty row in the sheet
#             next_row = sheet.max_row + 1

#             # Save the data
#             sheet.cell(row=next_row, column=1, value=name)
#             sheet.cell(row=next_row, column=2, value=website)
#             sheet.cell(row=next_row, column=3, value=phone)
#             sheet.cell(row=next_row, column=4, value=address)
#             sheet.cell(row=next_row, column=5, value=" ".join(categories))
#             sheet.cell(row=next_row, column=6, value=description)
#             sheet.cell(row=next_row, column=7, value=check_av_integration_needs_with_gpt(name + "  categories in which the company is engaged: " + " ".join(categories) + " some description about the company: " + description))
#             # sheet.cell(row=next_row, column=7, value=res1)
#             # sheet.cell(row=next_row, column=8, value=res2)

#             # Save the workbook
#             workbook.save("out.xlsx")
#         except Exception as e:
#             print(f"Error saving data to Excel: {e}")
import openpyxl

df = openpyxl.load_workbook("out.xlsx")
sheet = df.active
for row in sheet.iter_rows():
    i = row[0].value
    site = row[1].value
    phone_number = row[2].value
    adress = row[3].value
    categories = row[4].value
    about_little = row[5].value
    site_text = row[6].value

df.close()
