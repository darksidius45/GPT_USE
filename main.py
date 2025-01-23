from googlesearch import search
import re
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.options import Options
import time
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import WebDriverException
import openpyxl
import openai
import json
import subprocess
import requests
import random
from bs4 import BeautifulSoup
import zipfile

openai.api_key = 'not now baby'


output_file = r"C:\Users\prive_bodbgna\Downloads\comp_an.xlsx"
file_pars = "C:/Users/prive_bodbgna/Downloads/pars_me.txt"
plugin_file = "proxy_auth_plugin.zip"



PROXY_PORT = 3000
PROXY_USER = "ay30Uc"
PROXY_PASSWORD = "WTPKvo6e97"


manifest_json = """
{
        "version": "1.0.0",
        "manifest_version": 2,
        "name": "Chrome Proxy",
        "permissions": [
            "proxy",
            "tabs",
            "unlimitedStorage",
            "storage",
            "<all_urls>",
            "webRequest",
            "webRequestBlocking"
        ],
        "background": {
            "scripts": ["background.js"]
        }
}
"""

background_js_template = """
    var config = {
        mode: "fixed_servers",
        rules: {
            singleProxy: {
                scheme: "http",
                host: "%s",
                port: parseInt("%s")
            },
            bypassList: ["localhost"]
        }
    };

    chrome.proxy.settings.set({value: config, scope: "regular"}, function() {});

    function callbackFn(details) {
        return {
            authCredentials: {
                username: "%s",
                password: "%s"
            }
        };
    }

    chrome.webRequest.onAuthRequired.addListener(
        callbackFn,
        {urls: ["<all_urls>"]},
        ['blocking']
    );
    """

def get_random_proxy():
    proxies = [
        "194.32.229.90",
        "46.8.106.125",
        "109.248.48.163",
        "46.8.23.25",
        "5.183.130.240",
        "188.130.219.131",
        "95.182.124.146",
        "46.8.22.68",
        "46.8.16.52",
        "188.130.128.208",
        "46.8.106.142",
        "109.248.14.190",
        "188.130.136.182",
        "109.248.205.76",
        "109.248.167.155",
        "46.8.223.114",
        "45.15.72.93",
        "31.40.203.234",
        "46.8.23.61"
        "188.130.142.88",
        "109.248.130.150",
        "109.248.143.191",
        "188.130.129.164",
        "46.8.51.4",
        "94.198.190.78",
        "188.130.137.242",
        "188.130.143.89",
        "188.130.142.167",
        "109.248.13.204",
        "46.8.23.213",
        "194.32.229.90",
        "46.106.6.125",
        "109.248.48.163",
        "46.8.23.25",
        "5.183.130.240",
        "188.130.219.131",
        "95.182.124.146",
        "46.8.22.68",
        "46.8.16.52",
        "188.130.128.208",
        "46.8.108.142",
        "109.248.14.190"
    ]
    return random.choice(proxies)

def get_company_website(name, exhibitor_url):
    try:
        background_js = background_js_template % (get_random_proxy(), PROXY_PORT, PROXY_USER, PROXY_PASSWORD)
        with zipfile.ZipFile(plugin_file, "w") as zp:
            zp.writestr('manifest.json', manifest_json)
            zp.writestr('background.js', background_js)
        chrome_options = Options()
        chrome_options.add_argument("--headless")
        chrome_options.add_argument(f'--proxy-server=http://{proxy}')
        chrome_options.add_extension(plugin_file)  
        chrome_options.add_argument("user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/114.0.5735.110 Safari/537.36")
        service = Service("C:/webdrivers/chromedriver.exe")
        driver = webdriver.Chrome(service=service, options=chrome_options)
        try:
            driver.get(exhibitor_url)
            element = WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.CLASS_NAME, "exhibitor-details-contact-us-links"))
            )
            element_text = element.text
            driver.quit()
            return element_text
        except WebDriverException as e:
            raise e
        
    
    except Exception as e:
        query = f"{name} site"
        proxy = get_random_proxy()
        proxies = {
            "http": f"http://{proxy}",
            "https": f"http://{proxy}",
        }
    
        try:
            for result in search(query, num_results=10):
                if all(exclusion not in result for exclusion in ['youtube.com', 'facebook.com', 'wikipedia.org', 'linkedin.com']):
                    return result
        except requests.exceptions.HTTPError as e:
            print(f"HTTP error occurred: {e}")
        return None

def check_av_integration_needs_with_gpt(site):
    background_js = background_js_template % (get_random_proxy(), PROXY_PORT, PROXY_USER, PROXY_PASSWORD)
    with zipfile.ZipFile(plugin_file, "w") as zp:
        zp.writestr('manifest.json', manifest_json)
        zp.writestr('background.js', background_js)
    proxy = get_random_proxy()
    chrome_options = Options()
    chrome_options.add_argument("--headless")
    chrome_options.add_extension(plugin_file)  
    chrome_options.add_argument(f'--proxy-server=http://{proxy}')
    chrome_options.add_argument("user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/114.0.5735.110 Safari/537.36")
    service = Service("C:/webdrivers/chromedriver.exe")  
    driver = webdriver.Chrome(service=service, options=chrome_options)

    try:
        driver.get(site)
        time.sleep(3)
        text_html = driver.page_source
        soup = BeautifulSoup(text_html, "html.parser")
        text = soup.get_text(separator=" ", strip=True)
    except:
        driver.quit()
        return "ERROR"
    prompt = f" can you say if this company may need service of installing video walls on practice in nearby future, pls, be very strict (all reasons must be significant(for examle shoe ompany dont need it))(first i need one word yes or no and then reason){text}"
    try:
        curl_command = [
    
        "curl",
        "https://api.proxyapi.ru/openai/v1/chat/completions",
        "-H", "Content-Type: application/json",
        "-H", "Authorization: Bearer sk-UmoxJZ8wJm1DEmcSrwP3Iu9Bk4TGZJ0h", 
        "-d", json.dumps({
            "model": "gpt-4o-mini",
            "messages": [{"role": "user", "content": prompt}],
            "stream": False 
        })
        ]

    
        result = subprocess.run(curl_command, capture_output=True, text=True, encoding='utf-8')


        if result.returncode != 0:
            print("Error executing curl command:", result.stderr)
            driver.quit()
            return "rrrrr"

    
        response = json.loads(result.stdout)
        return response['choices'][0]['message']['content']
    except Exception as e:
        print(f"Error checking site {site}: {e}")
        driver.quit()
        return "eror"



counter = 552


names = set()
with open(file_pars, "r", encoding='latin-1') as f:
    content = f.read()


start_tag = '<div class="exhibitor-summary">'
end_tag = '</div>'
pattern = re.compile(f'{re.escape(start_tag)}(.*?){re.escape(end_tag)}', re.DOTALL)
matches = pattern.findall(content)
print(len(matches))
exhibitors = {}


for match in matches:
    name_pattern = re.compile(r'<h3 class="exhibitor-name wrap-word">(.*?)</h3>')
    name_match = name_pattern.search(match)
    name = name_match.group(1) if name_match else None

    website_pattern = re.compile(r'<a href="(.*?)" data-dtm="')
    website_match = website_pattern.search(match)
    website = website_match.group(1).strip() if website_match else None

    if name and website:
        exhibitors[name] = website
        for key in exhibitors:
        wb = openpyxl.load_workbook(output_file)
        sheet = wb.active
        exhibitors[key] = get_company_website(key, exhibitors[key])
        print(f"Company website for {key}: {exhibitors[key]}", end="\n")
        sheet.cell(row=counter, column=1, value=key)
        sheet.cell(row=counter, column=2, value=exhibitors[key])
        sheet.cell(row=counter, column=3, value = check_av_integration_needs_with_gpt(exhibitors[key]))
        counter += 1
        wb.save(output_file)