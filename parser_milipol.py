import openpyxl
import json
import subprocess
import tempfile
import os
from bs4 import BeautifulSoup

output_file = r"C:\Users\prive\Desktop\prog\rgb\GPT_USE\output_files\output_milipol.xlsx"
file_pars = r"C:\Users\prive\Desktop\prog\rgb\GPT_USE\pars_files\pars_milipol.txt"
company_pars = r"C:\Users\prive\Desktop\prog\rgb\GPT_USE\pars_files\company_names.xlsx"

company_names = set()

def normalize_name(name):
    normalized_name = name.lower()  # Convert to lowercase
    normalized_name = normalized_name.replace('–', '-')  # Replace long dashes with regular dashes
    normalized_name = normalized_name.replace('—', '-')  # Replace em dashes with regular dashes  # Replace hyphens with spaces
    normalized_name = ' '.join(normalized_name.split())  # Remove extra spaces
    return normalized_name

def gemini_flash(text):
    # Create a temporary file to store the JSON data
    with tempfile.NamedTemporaryFile(delete=False, mode='w', suffix='.json') as temp_file:
        json_data = {
            "contents": [{"role": "user", "parts": [{"text": text}]}]
        }
        json.dump(json_data, temp_file)
        temp_file.flush()  # Ensure the data is written

        # Get the name of the temporary file
        temp_file_name = temp_file.name

    try:
        # Simplified curl command using the temporary file
        curl_command = [
            "curl",
            "https://api.proxyapi.ru/google/v1/models/gemini-1.5-flash:generateContent",
            "-H", "Content-Type: application/json",
            "-H", "Authorization: Bearer sk-UmoxJZ8wJm1DEmcSrwP3Iu9Bk4TGZJ0h",
            "-d", f"@{temp_file_name}"  # Use the temp file
        ]

        result = subprocess.run(curl_command, capture_output=True, text=True, encoding='utf-8')
        if result.returncode == 0:
            response = json.loads(result.stdout)
            text_response = response['candidates'][0]['content']['parts'][0]['text']
            print(text_response)
            return text_response
        else:
            print("Error executing curl command:", result.stderr)
            return None
    finally:
        # Clean up the temporary file
        if os.path.exists(temp_file_name):
            os.remove(temp_file_name)



df = openpyxl.load_workbook(output_file)  # Load the existing workbook
sheet = df.active

# sheet.delete_cols(1)


# with open(file_pars, 'r', encoding='utf-8') as f:  # Specify encoding to avoid UnicodeDecodeError
#     content = f.read()

#     soup = BeautifulSoup(content, 'html.parser')
#     i = 1
#     for h3_tag in soup.find_all('h3', class_='CatalogCardLine-title textAccent'):
#         company_name = normalize_name(h3_tag.get_text(strip=True))
#         sheet[f"A{i}"].value = company_name
#         i += 1
#         df.save(output_file)
#         print(company_name)

# company_data = openpyxl.load_workbook(company_pars)
# company_sheet = company_data.active
# i = 1
# for row in company_sheet.iter_rows(min_row=2, values_only=True):
#     company_name = row[0]
#     normalized_name = normalize_name(company_name)
#     sheet[f"B{i}"].value = normalized_name 
#     i += 1
# df.save(output_file)

company_names1 = []
company_names2 = []

for row in sheet.iter_rows(min_row=1, values_only=True):
    company_name1 = row[0]
    company_name2 = row[1]
    company_names1.append(company_name1)
    company_names2.append(company_name2)

print(gemini_flash(f"find the common companies in the following lists, altho write down, companies with very similar names, but not the same if it is possible for them to be the same company (please note that the names may be different due to the fact that there are words with the designation of countries, the words co, ltd and other words that do not relate to the name itself.): the first list is {company_names1} and the second list is {company_names2}"))


df.close()  # Save changes to the output file

