import openpyxl
import json
import subprocess
import tempfile
import os

output_file = r"C:\Users\prive\Desktop\prog\rgb\GPT_USE\output_files\output_siae.xlsx"
file_pars = r"C:\Users\prive\Desktop\prog\rgb\GPT_USE\pars_files\pars_siae.txt"
company_pars = r"C:\Users\prive\Desktop\prog\rgb\GPT_USE\pars_files\company_names.xlsx"

company_names = set()


def normalize_name(name):
    if name is None:
        return ""  # Return empty string for None values
    normalized_name = name.lower()  # Convert to lowercase
    normalized_name = normalized_name.replace(
        "–", "-"
    )  # Replace long dashes with regular dashes
    normalized_name = normalized_name.replace(
        "—", "-"
    )  # Replace em dashes with regular dashes  # Replace hyphens with spaces
    normalized_name = " ".join(normalized_name.split())
    # Remove extra spaces and common company suffixes/words
    normalized_name = (
        normalized_name.replace(" inc", "").replace(" llc", "").replace(" ltd", "")
    )
    normalized_name = (
        normalized_name.replace(" gmbh", "").replace(" srl", "").replace(" sa", "")
    )
    normalized_name = normalized_name.replace(" technologies", "").replace(
        " technology", ""
    )
    normalized_name = normalized_name.replace(" international", "").replace(" corp", "")
    normalized_name = normalized_name.replace(" co", "").replace(" company", "")
    normalized_name = normalized_name.replace(" group", "").replace(" holding", "")
    normalized_name = normalized_name.replace(" systems", "").replace(" solutions", "")
    normalized_name = normalized_name.replace(" industries", "").replace(
        " industrial", ""
    )
    normalized_name = (
        normalized_name.replace(" ag", "").replace(" nv", "").replace(" bv", "")
    )
    normalized_name = normalized_name.replace(" limited", "").replace(
        " incorporated", ""
    )
    normalized_name = " ".join(
        normalized_name.split()
    )  # Remove any remaining extra spaces
    return normalized_name


def gemini_flash(text):
    # Create a temporary file to store the JSON data
    with tempfile.NamedTemporaryFile(
        delete=False, mode="w", suffix=".json"
    ) as temp_file:
        json_data = {"contents": [{"role": "user", "parts": [{"text": text}]}]}
        json.dump(json_data, temp_file)
        temp_file.flush()  # Ensure the data is written

        # Get the name of the temporary file
        temp_file_name = temp_file.name

    try:
        # Simplified curl command using the temporary file
        curl_command = [
            "curl",
            "https://api.proxyapi.ru/google/v1/models/gemini-1.5-flash:generateContent",
            "-H",
            "Content-Type: application/json",
            "-H",
            "Authorization: Bearer sk-UmoxJZ8wJm1DEmcSrwP3Iu9Bk4TGZJ0h",
            "-d",
            f"@{temp_file_name}",  # Use the temp file
        ]

        result = subprocess.run(
            curl_command, capture_output=True, text=True, encoding="utf-8"
        )
        if result.returncode == 0:
            response = json.loads(result.stdout)
            text_response = response["candidates"][0]["content"]["parts"][0]["text"]
            print(text_response)
            return text_response
        else:
            print("Error executing curl command:", result.stderr)
            return None
    finally:
        # Clean up the temporary file
        if os.path.exists(temp_file_name):
            os.remove(temp_file_name)


df = openpyxl.load_workbook(output_file)  # Load the existing workbook
sheet = df.active


# with open(file_pars, 'r', encoding='utf-8') as f:  # Specify encoding to avoid UnicodeDecodeError
#     content = f.read()

#     soup = BeautifulSoup(content, 'html.parser')
#     for div_tag in soup.find_all('div', class_='card-nomExposant'):
#         company_name = normalize_name(div_tag.text.strip())
#         company_names.add(company_name)
#         sheet.append([company_name])
#         df.save(output_file)
#         print(company_name)

# with open(company_pars, 'r', encoding='utf-8') as f:
#     company_data = openpyxl.load_workbook(company_pars)
#     company_sheet = company_data.active
#     i = 1
#     for row in company_sheet.iter_rows(min_row=2, values_only=True):
#         company_name = row[0]
#         normalized_name = normalize_name(company_name)
#         sheet[f"B{i}"].value = normalized_name
#         i += 1
company_names1 = []
company_names2 = set()

for row in sheet.iter_rows(min_row=1, values_only=True):
    company_name1 = normalize_name(row[0])
    company_name2 = normalize_name(row[1])
    if company_name1:  # Only add non-empty names
        company_names1.append(company_name1)
    if company_name2:  # Only add non-empty names
        company_names2.add(company_name2)


print(
    gemini_flash(
        f"find the common companies in the following lists, altho write down, companies with very similar names, but not the same if it is possible for them to be the same company (please note that the names may be different due to the fact that there are words with the designation of countries, the words co, ltd and other words that do not relate to the name itself., for example, if the company is called 'company' and the other is 'company co', it is the same company(example:extron electronics europe inc. is the same as extron)): the first list is {" ".join(company_names1)} and the second list is {" ".join(company_names2)}"
    )
)


df.close()  # Save changes to the output file
