from googlesearch import search
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import WebDriverException
import openpyxl
import json
import subprocess
import random
from bs4 import BeautifulSoup
import zipfile
from urllib.parse import urlparse, urljoin
import tempfile
import os

distributor = (
    "The integrator is the one who produces the work. His task is to design to coordinate the operation of different systems - for example, in the control room, the operator should see the cameras of the surveillance system and the air conditioning system components. "
    "Both the sales work and the physical installation of the distributor - wholesaler are usually distinguishable by the large number of brands they sell. "
    "Integrator and Distributor often merge in business because it is more convenient for the integrator to integrate the goods available in his warehouse - that is, to use the distributor's portfolio. "
    "And it is convenient for a distributor to have an integration department because it is like a retail store in a wholesale warehouse."
    "Manufacturers is distinguishable by the absence of other brands on the site, but it exhibits models of similar equipment with similar part numbers."
)

av_integrator = (
    "AV Integration for Video Systems: Explanation\n\n"
    "AV Integration for Video Systems focuses on the design, installation, and management of video-centric technologies to create seamless visual communication and presentation solutions. This process involves combining hardware, software, and network infrastructure to ensure high-quality video delivery, display, and control in various environments. Unlike full AV integration, which includes audio components, this explanation is tailored specifically to video systems.\n\n"
    "Key Components of Video-Centric AV Integration:\n"
    "Displays:\n"
    "- LED Walls: Large-scale, high-resolution displays for impactful visuals in events, lobbies, or control rooms.\n"
    "- LCD/LED Screens: Flat-panel displays for boardrooms, classrooms, or digital signage.\n"
    "- Projectors and Screens: Ideal for large audiences in auditoriums, theaters, or conference halls.\n\n"
    "Video Sources:\n"
    "- Media Players: Devices for playing pre-recorded video content.\n"
    "- Streaming Devices: Tools for live video streaming or accessing online content.\n"
    "- Cameras: Video conferencing cameras, PTZ (pan-tilt-zoom) cameras, or production cameras for live events.\n\n"
    "Signal Management:\n"
    "- Switchers and Matrix Systems: Devices to route and manage multiple video signals to various displays.\n"
    "- Extenders: Tools to transmit video signals over long distances (HDMI, SDI, or IP-based extenders).\n"
    "- Converters: Devices to convert video signals between different formats (e.g., HDMI to SDI).\n\n"
    "Control Systems:\n"
    "- Touch Panels or Remotes: User-friendly interfaces to control video sources, displays, and routing.\n"
    "- Automation Systems: Integration with room controls to automate video playback, display settings, or lighting."
)
def get_random_proxy():
    proxies = [
        "45.11.21.124",
        "213.226.101.147",
        "94.158.190.16",
        "109.248.128.64",
        "188.130.218.160",
        "46.8.15.51",
        "46.8.23.126",
        "188.130.129.246",
        "188.130.128.66",
        "46.8.106.76",
        "46.8.111.80",
        "109.248.142.228",
        "46.8.222.147",
        "109.248.204.170",
    ]
    return random.choice(proxies)


PROXY_PORT = 5500
PROXY_USER = "5vWk59"
PROXY_PASS = "tXQt5fnHW9"


def plugin(file_name):
    proxy = get_random_proxy()

    manifest_json = """
    {
        "version": "1.0.0",
        "manifest_version": 2,
        "name": "Chrome Proxy",
        "permissions": [
            "proxy",
            "tabs",
            "unlimitedStorage",
            "storage",
            "<all_urls>",
            "webRequest",
            "webRequestBlocking"
        ],
        "background": {
            "scripts": ["background.js"]
        },
        "minimum_chrome_version":"22.0.0"
    }
    """

    background_js = """
    var config = {
            mode: "fixed_servers",
            rules: {
            singleProxy: {
                scheme: "http",
                host: "%s",
                port: parseInt(%s)
            },
            bypassList: ["localhost"]
            }
        };

    chrome.proxy.settings.set({value: config, scope: "regular"}, function() {});

    function callbackFn(details) {
        return {
            authCredentials: {
                username: "%s",
                password: "%s"
            }
        };
    }

    chrome.webRequest.onAuthRequired.addListener(
                callbackFn,
                {urls: ["<all_urls>"]},
                ['blocking']
    );
    """ % (
        proxy,
        PROXY_PORT,
        PROXY_USER,
        PROXY_PASS,
    )

    with zipfile.ZipFile(file_name, "w") as zp:
        zp.writestr("manifest.json", manifest_json)
        zp.writestr("background.js", background_js)
    return proxy


def create_driver():
    chrome_options = Options()
    chrome_options.add_argument(
        "user-agent=Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/605.1.15 (KHTML, like Gecko) Version/14.0.3 Safari/605.1.15"
    )

    # Disable images
    chrome_prefs = {
        "profile.managed_default_content_settings.images": 2,
        "profile.default_content_setting_values.notifications": 2,
    }
    chrome_options.add_experimental_option("prefs", chrome_prefs)

    pluginfile = "proxy_auth_plugin.zip"
    chrome_options.add_extension(pluginfile)
    print(plugin(pluginfile))

    # chrome_options.add_argument('--headless')
    chrome_options.add_argument("--no-sandbox")
    chrome_options.add_argument("--disable-dev-shm-usage")

    service = Service("C:/webdrivers/chromedriver.exe")
    driver = webdriver.Chrome(service=service, options=chrome_options)

    return driver


def get_site_text(website):
    keywords = [
        "about",
        "product",
        "solution",  # English
        "über",
        "produkt",
        "lösung",  # German
        "acerca de",
        "producto",
        "solución",  # Spanish
        "à propos",
        "produit",
        "solution",  # French
        "circa",
        "prodotto",
        "soluzione",  # Italian
        "sobre",
        "produto",
        "solução",  # Portuguese
        "关于",
        "产品",
        "解决方案",  # Chinese
        "о",
        "продукт",
        "решение",  # Russian
        "について",
        "製品",
        "ソリューション",  # Japanese
        "hakkında",
        "ürün",
        "çözüm",  # Turkish
        "حول",
        "منتج",
        "حل",  # Arabic
    ]
    all_texts = []

    attempts = 0
    max_attempts = 5
    success = False
    driver = create_driver()

    try:
        while attempts < max_attempts and not success:
            try:
                driver.get(website)
                # Wait for body element explicitly
                WebDriverWait(driver, 10).until(
                    EC.presence_of_element_located((By.TAG_NAME, "body"))
                )
                success = True

                page_source = driver.page_source
                soup = BeautifulSoup(page_source, "html.parser")
                all_texts.append(soup.get_text(separator=" ", strip=True))

                original_domain = urlparse(website).netloc
                print("original_domain ", original_domain)
                hrefs_set = set()
                clickable_elements = set()

                for element in soup.find_all(["a", "button", "input", "area"]):
                    if element.has_attr("href"):
                        href = element["href"]
                        if href not in hrefs_set and any(
                            keyword in href.lower() for keyword in keywords
                        ):
                            clickable_elements.add(element["href"])

                print("clickable_elements ", clickable_elements)

                for element in clickable_elements:
                    new_link = urljoin(website, element)
                    driver.get(new_link)
                    # Wait for body element explicitly
                    WebDriverWait(driver, 10).until(
                        EC.presence_of_element_located((By.TAG_NAME, "body"))
                    )
                    page_source = driver.page_source
                    soup = BeautifulSoup(page_source, "html.parser")
                    # Get text content
                    text = soup.get_text(separator=" ", strip=True)

                    # Find and remove cookie-related blocks/sections
                    cookie_sections = soup.find_all(
                        ["div", "section", "footer", "aside"],
                        class_=lambda x: x
                        and any(
                            term in str(x).lower()
                            for term in [
                                "cookie",
                                "consent",
                                "gdpr",
                                "privacy-notice",
                                "privacy-banner",
                                "cookie-banner",
                            ]
                        ),
                    )

                    for section in cookie_sections:
                        section.decompose()

                    # Get clean text after removing cookie sections
                    clean_text = soup.get_text(separator="\n", strip=True)

                    # Split into lines and remove duplicates while preserving order
                    lines = clean_text.split("\n")

                    # Remove similar lines from both current text and all_texts
                    seen = set()
                    unique_lines = []

                    # First check against existing all_texts content
                    existing_text = " ".join(all_texts)
                    existing_lines = existing_text.split("\n")
                    for existing_line in existing_lines:
                        existing_line = existing_line.strip()
                        if existing_line:
                            seen.add(existing_line)

                    # Then process new lines
                    for line in lines:
                        line = line.strip()
                        if line and line not in seen:
                            seen.add(line)
                            unique_lines.append(line)

                    text = " ".join(unique_lines)
                    all_texts.append(soup.get_text(separator=" ", strip=True))

            except WebDriverException:
                attempts += 1
                print(f"Attempt {attempts} failed. Retrying with new proxy...")
                driver.quit()
                driver = create_driver()

                if attempts == max_attempts:
                    print("Max attempts reached. Could not access website.")
                    return None

    except Exception as e:
        print(f"An error occurred: {e}")
        return None

    finally:
        driver.quit()

    return " ".join(all_texts)


def gemini_flash(text):
    # Create a temporary file to store the JSON data
    with tempfile.NamedTemporaryFile(
        delete=False, mode="w", suffix=".json"
    ) as temp_file:
        json_data = {"contents": [{"role": "user", "parts": [{"text": text}]}]}
        json.dump(json_data, temp_file)
        temp_file.flush()  # Ensure the data is written

        # Get the name of the temporary file
        temp_file_name = temp_file.name

    try:
        # Simplified curl command using the temporary file
        curl_command = [
            "curl",
            "https://api.proxyapi.ru/google/v1/models/gemini-1.5-flash:generateContent",
            "-H",
            "Content-Type: application/json",
            "-H",
            "Authorization: Bearer sk-UmoxJZ8wJm1DEmcSrwP3Iu9Bk4TGZJ0h",
            "-d",
            f"@{temp_file_name}",  # Use the temp file
        ]

        result = subprocess.run(
            curl_command, capture_output=True, text=True, encoding="utf-8"
        )
        if result.returncode == 0:
            try:
                response = json.loads(result.stdout)
                text_response = response["candidates"][0]["content"]["parts"][0]["text"]
                print(text_response)
                return text_response
            except (KeyError, IndexError) as e:
                print(f"Error parsing response: {e}")
                print(f"Response content: {result.stdout}")
                return None
        else:
            print("Error executing curl command:", result.stderr)
            return None
    finally:
        # Clean up the temporary file
        if os.path.exists(temp_file_name):
            os.remove(temp_file_name)


df = openpyxl.load_workbook(r"C:\Users\prive\Desktop\prog\rgb\GPT_USE\output_files\out.xlsx")
write = openpyxl.load_workbook(r"C:\Users\prive\Desktop\prog\rgb\GPT_USE\output_files\new_output3.xlsx")




sheet = df.active
write_sheet = write.active


for row in sheet.iter_rows(min_row=1415):
    name = row[0].value
    site = row[1].value
    phone_number = row[2].value
    adress = row[3].value
    categories = row[4].value
    about_little = row[5].value
    site_text = get_site_text(site)
    if site_text != "":
        site_info = gemini_flash(
        f"retelling the text {site_text}.Say if it is distributor, integrator or manufacturer according to the text: {distributor}. Focus on what the company does (manufactures or resells) and tell us about their products and services."
        )
        answer = gemini_flash(
            f"tell me if this company is connected with my defenition of av integrations(AV integrations: {av_integrator})(i need 1 world yes or no in the begging and only then why) company info - categroies:{categories} , about them{about_little}, their site:{site_info}. Remeber i need only video intergrations"
        )
    else:
        # answer = "error"
        site_info = "error"
    
    cleaned_values = []
    for value in [name, site, phone_number, adress, categories, about_little, site_text]:
        if value is None:
            cleaned_values.append("")
        else:
            # Remove or replace problematic characters
            cleaned_value = str(value).encode('ascii', 'ignore').decode()
            cleaned_values.append(cleaned_value)
            
    write_sheet.append(cleaned_values)

    write.save(r"C:\Users\prive\Desktop\prog\rgb\GPT_USE\output_files\new_output3.xlsx")
df.close()
